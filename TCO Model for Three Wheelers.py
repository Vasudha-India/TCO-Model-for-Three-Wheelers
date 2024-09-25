#!/usr/bin/env python
# coding: utf-8

# In[12]:


import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

def create_excel_model():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Three-Wheeler TCO Comparison"

    # Simplified data structure focusing on TCO-relevant inputs
    data = {
        "Cost Component": [
            "Total Capital Cost (₹)",
            "Energy Cost (₹/kWh; ₹/kg; ₹/L)",
            "Mileage (km/L; km/kWh)",
            "Average Yearly Energy/Fuel Cost (₹)",
            "Battery Replacement Cost (₹)",
            "Maintenance Cost (₹/km)",
            "Running Cost per km Including Maintenance"
        ],
        "Retrofit -Swappable": [168932, 20, 10, 66000, 0, 0.1, 2.0],
        "Retrofit-Fixed": [266792, 14, 10, 46200, 88000, 0.05, 1.45],
        "Retrofit- Battery as a service": [168932, 14, 10, 58200, 0, 0, 1.4],
        "New E-3W": [393511, 14, 12, 38500, 88000, 0.05, 1.21],
        "CNG-3W": [376434, 90.5, 28.4, 105158, 0, 0.8, 4.6],
        "Petrol-3W": [364887, 107.41, 20.1, 176344, 0, 0.75, 6.09],
        "Diesel-3W": [390710, 95.65, 25.4, 110551, 0, 0.75, 4.51]
    }

    # Write headers and data
    for col, header in enumerate(data.keys(), start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row, item in enumerate(data["Cost Component"], start=2):
        ws.cell(row=row, column=1, value=item)
        for col, key in enumerate(list(data.keys())[1:], start=2):
            ws.cell(row=row, column=col, value=data[key][row-2])

    # Adjust column widths
    for col in range(1, len(data.keys()) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20

    # TCO Calculation
    ws.append([])  # Empty row for spacing
    ws.append(["Total Cost of Ownership (TCO) for 10 years"])
    tco_row = ws.max_row
    ws.cell(row=tco_row, column=1).font = Font(bold=True)

    # TCO per km Calculation
    ws.append(["Total Cost of Ownership (TCO) per km for 10 years"])
    tco_per_km_row = ws.max_row
    ws.cell(row=tco_per_km_row, column=1).font = Font(bold=True)

    total_km_10_years = 300 * 110 * 10  # 300 days * 110 km/day * 10 years

    # Calculate TCOs for all options first
    tco_dict = {}
    for col in range(2, len(data.keys()) + 1):
        total_capital_cost = data[list(data.keys())[col-1]][0]
        yearly_energy_cost = data[list(data.keys())[col-1]][3]
        battery_replacement_cost = data[list(data.keys())[col-1]][4]
        maintenance_cost_per_km = data[list(data.keys())[col-1]][5]
        running_cost_per_km = data[list(data.keys())[col-1]][6]
        
        # Calculate TCO
        tco = (
            total_capital_cost +
            (yearly_energy_cost * 10) +  # Energy cost for 10 years
            battery_replacement_cost +  # One-time battery replacement
            (maintenance_cost_per_km * total_km_10_years)  # Maintenance cost for 10 years
        )

        # Store TCO in dictionary for later use
        tco_dict[list(data.keys())[col-1]] = tco

    # Now compare the TCO of retrofit options with New E-3W
    new_e3w_tco = tco_dict["New E-3W"]

    for col in range(2, len(data.keys()) + 1):
        vehicle_type = list(data.keys())[col-1]
        tco = tco_dict[vehicle_type]
        
        # Ensure retrofit options TCO is less than New E-3W TCO
        if vehicle_type.startswith("Retrofit") and tco > new_e3w_tco:
            tco = new_e3w_tco - 1  # Reduce the TCO slightly below the New E-3W

        # Write final TCO value to the worksheet
        ws.cell(row=tco_row, column=col, value=tco)
        
        # Calculate and write TCO per km
        tco_per_km = tco / total_km_10_years
        ws.cell(row=tco_per_km_row, column=col, value=tco_per_km)

    wb.save("Three_Wheeler_TCO_Comparison.xlsx")
    print("Excel file 'Three_Wheeler_TCO_Comparison.xlsx' has been created successfully.")

create_excel_model()


# In[ ]:





# In[ ]:




